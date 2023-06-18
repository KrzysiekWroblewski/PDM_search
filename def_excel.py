import openpyxl
from excel_raport_generator import Report
from List_sorting import Sort

# open_file = "sample.xlsx"


def read_excel_BOM(open_file):

    xl_part_column = 7
    csv_sep = "|"

    wrkbk = openpyxl.load_workbook(open_file)
    sh = wrkbk.active

    search_Path_sld_prt = ""
    search_Path_sld_prt_revision = ""
    number_of_files_merged = 0

    list_of_records = []

    for j in range(6, sh.max_row+1):
        cell_obj = sh.cell(row=j, column=xl_part_column)
        cell_revision = sh.cell(row=j, column=xl_part_column+1)

        # skip line if par is mirror
        if "lustro" in sh.cell(row=j, column=xl_part_column).value.lower():
            continue

        # Checking if  are some missing drawings
        for i in range(3, 6):
            if sh.cell(row=j, column=i).value == None:
                drawings = False
                break
            else:
                drawings = True

        if drawings == False:
            list_of_records.append(
                [sh.cell(row=j, column=1).value, sh.cell(row=j, column=6).value, sh.cell(row=j, column=xl_part_column).value, sh.cell(row=j, column=xl_part_column+1).value, sh.cell(row=j, column=xl_part_column+2).value])

        if drawings == False:
            search_Path_sld_prt = search_Path_sld_prt + \
                str(cell_obj.value) + csv_sep

            number_of_files_merged = number_of_files_merged + 1

        if drawings == False:
            search_Path_sld_prt_revision = search_Path_sld_prt_revision + \
                str(cell_obj.value) + "-" + str(cell_revision.value) + csv_sep

    try:
        if search_Path_sld_prt[-1] == "|":
            search_Path_sld_prt = search_Path_sld_prt[:-1]
            search_Path_sld_prt_revision = search_Path_sld_prt_revision[:-1]
    except:
        pass

    search_Path_sld_prt = search_Path_sld_prt.encode(
        'utf-8').decode('ascii', 'ignore')

    search_Path_sld_prt_revision = search_Path_sld_prt_revision.encode(
        'utf-8').decode('ascii', 'ignore')

    print(search_Path_sld_prt)
    print(search_Path_sld_prt_revision)
    print(number_of_files_merged)

    name_of_file = open_file.replace("/", "_")
    name_of_file = name_of_file[2:-5]

    Report.generate_report_txt(Sort.list_sort_by_index(
        list_of_records, 1), name_of_file, search_Path_sld_prt.encode("utf-8"), search_Path_sld_prt_revision.encode("utf-8"))

    return (number_of_files_merged, search_Path_sld_prt, search_Path_sld_prt_revision)
