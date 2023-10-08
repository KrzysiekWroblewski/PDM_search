
from PyPDF2 import PdfWriter, PdfReader
import os
from regex import B
from gui import GUI
import openpyxl
from EXCEL_report_generator import Report
import os
from EXCEL_methods import excel
from gui import GUI
from openpyxl import Workbook
from EXCEL_report_generator import Report
import datetime
from sql import SQL
import sys
import items


def Generate_QR_to_png(text, QR_png_path, size=50, border_size=1):
    import qrcode
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=size,
        border=border_size
    )
    qr.add_data(text)
    qr.make(fit=True)

    img = qr.make_image(fill_color="white", back_color="black")
    img.save(QR_png_path)

    return QR_png_path


def Convert_QR_PNG_to_PDF(QR_png_path, QR_pdf_path, position_y=100, position_x=100):
    from reportlab.pdfgen import canvas

    # Set the position and size of the image
    x = int(position_x) - 139.5  # 352.4 - 1091  # X-coordinate of the image
    y = 31

    position_x = 56  # Width of the image
    position_y = 56  # Height of the image

    new_pdf = canvas.Canvas(QR_pdf_path)
    new_pdf.drawImage(QR_png_path, x, y, position_x, position_y)
    new_pdf.save()

    return QR_pdf_path


def copy_file(source_directory, destination_directory):
    import shutil

    shutil.copy2(source_directory, destination_directory)


def date_stamp() -> str:
    date_stamp = datetime.datetime.now()
    date_stamp_formated = date_stamp.strftime("%Y-%m-%d %H-%M")
    return date_stamp_formated


def files_from_dir():
    directory = GUI.select_folder()  # Replace with the path to your directory

    # Get all file names in the directory
    file_names = os.listdir(directory)
    list = []
    # Print the file namess
    for file_name in file_names:
        list.append(file_name)
    return list


def Copy_pdf_to_new_dir(oryginal_path, new_file_path):
    from PyPDF2 import PdfReader, PdfWriter
    from PyPDF2.generic import AnnotationBuilder

    # Fill the writer with the pages you want
    pdf_path = oryginal_path
    reader = PdfReader(pdf_path)
    page = reader.pages[0]
    writer = PdfWriter()
    writer.add_page(page)

    # read page dimensions
    box = reader.pages[0].mediabox
    page_width = box.width
    page_height = box.height

    # Write the annotated file to disk
    with open(new_file_path, "wb") as fp:
        writer.write(fp)

    return (page_width, page_height)


def create_pdf_with_text_box(text_box_pdf_path, height_pdf, width_pdf, line0="", line1="", line2="", line3="", line4=""):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    # Create a canvas object
    c = canvas.Canvas(text_box_pdf_path+".pdf", pagesize=letter)

    # Set the coordinates and size of the text box

    # Size of  the text box
    x = 172
    y = 53.8
    position_x = width_pdf - 311
    position_y = 32.4

    c.rect(position_x, position_y, x, y)

    # Add text inside the text box
    c.setFont("Helvetica", 9)
    c.drawString(position_x + 5, y + position_y -
                 10, f"1.ID:  {line0}")  # line1
    c.drawString(position_x + 5, y + position_y -
                 20, f"2.Ilosc: {line1}")  # line2
    c.drawString(position_x + 5, y + position_y - 30,
                 f"3.Nr zamowienia: {line2}")  # line3
    c.drawString(position_x + 5, y + position_y -
                 40, f"4.Data zamowienia: {line3}")  # line3
    c.drawString(position_x + 5, y + position_y -
                 50, f"5.Zamawiajacy: {line4}")  # line3
    # Save the canvas and close it
    c.save()
    return


def Insert_text_box_on_pdf(content_pdf_path, stamp_pdf_path, pdf_result_path, page_indices="ALL"):

    reader = PdfReader(content_pdf_path)
    if page_indices == "ALL":
        page_indices = list(range(0, len(reader.pages)))

    writer = PdfWriter()

    content_page = reader.pages[0]
    mediabox = content_page.mediabox

    # You need to load it again, as the last time it was overwritten
    reader_stamp = PdfReader(stamp_pdf_path)
    image_page = reader_stamp.pages[0]

    image_page.merge_page(content_page)
    image_page.mediabox = mediabox

    writer.add_page(image_page)

    with open(pdf_result_path, "wb") as fp:
        writer.write(fp)


def Make_QR_stamp_on_pdf(content_pdf, stamp_pdf, pdf_result, page_indices="ALL"):

    reader = PdfReader(content_pdf)
    if page_indices == "ALL":
        page_indices = list(range(0, len(reader.pages)))

    writer = PdfWriter()

    content_page = reader.pages[0]
    mediabox = content_page.mediabox

    # You need to load it again, as the last time it was overwritten
    reader_stamp = PdfReader(stamp_pdf)
    image_page = reader_stamp.pages[0]

    image_page.merge_page(content_page)
    image_page.mediabox = mediabox

    writer.add_page(image_page)

    with open(pdf_result, "wb") as fp:
        writer.write(fp)


def merge_pdfs(input_files, output_file):
    from PyPDF2 import PdfMerger

    merger = PdfMerger()

    # Open each input file and add it to the merger
    for file in input_files:
        merger.append(file)

    # Write the merged PDF to the output file
    merger.write(output_file)

    # Close the merger
    merger.close()


def Find_items_with_value_in_cell(list_of_items: list, searched_column: int, order_name):
    list_of_found_items = []
    for item in list_of_items:
        if item[searched_column].lower().replace(" ", "") == order_name.lower().replace(" ", ""):
            list_of_found_items.append(item)
            # print("i Found unordered item!")
        else:
            continue
    return list_of_found_items


def Find_column_index_by_string(lista, szukany_element):
    for indeks, element in enumerate(lista[0]):
        if element in szukany_element:
            return indeks
    return -1  # zwracamy -1, jeśli element nie został znaleziony


def Write_excel_BOM_to_list(open_file: str, start_row: int) -> list:
    """Def open active sheet of work book and write rows to lists

    Args:
        open_file (str): Path to file which is rewrite to list

    Returns:
        list[str]: strings in list are encoded to bytes
    """

    wrkbk = openpyxl.load_workbook(open_file, data_only=True)
    sheet_index = 2
    sh = wrkbk.worksheets[sheet_index - 1]

    list_of_records = []

    for i in range(start_row, sh.max_row+1):
        details_of_record = []
        for j in range(1, sh.max_column+1):
            details_of_record.append(
                str(sh.cell(row=i, column=j).value).replace("\n", ""))
        list_of_records.append(details_of_record)
    # print('list_of_records: ', list_of_records)

    return list_of_records


def EXCEL_orders(open_file: str, order_name):
    List_of_items = Write_excel_BOM_to_list(open_file, start_row=5)

    # find column with header "Numer zamówienia"
    header_title = "Numer zamówienia"
    column_index = Find_column_index_by_string(List_of_items, header_title)

    # find rows with None in column
    List_of_records = Find_items_with_value_in_cell(
        List_of_items, column_index, order_name)

    Lp_index = excel.Excel_Find_column_index_by_string_in_list(
        List_of_items, "Lp")
    Konstruktor_index = excel.Excel_Find_column_index_by_string_in_list(
        List_of_items, "Konstruktor")
    id_index = excel.Excel_Find_column_index_by_string_in_list(
        List_of_items, "ID. Części")
    Rewizja_index = excel.Excel_Find_column_index_by_string_in_list(
        List_of_items, "Rewizja")
    Opis_index = excel.Excel_Find_column_index_by_string_in_list(
        List_of_items, "Opis")
    Drawing_Number_index = excel.Excel_Find_column_index_by_string_in_list(
        List_of_items, "Nr. Części ")
    Material_index = excel.Excel_Find_column_index_by_string_in_list(
        List_of_items, "Materiał")
    Obrobka_index = excel.Excel_Find_column_index_by_string_in_list(
        List_of_items, "Obróbka")
    Uwagi_index = excel.Excel_Find_column_index_by_string_in_list(
        List_of_items, "Uwagi:")
    suma_ilosc_index = excel.Excel_Find_column_index_by_string_in_list(
        List_of_items, "Do Zamówienia")

    headers_indexes = [Lp_index, Konstruktor_index, id_index,  Rewizja_index,
                       Opis_index, Drawing_Number_index, Material_index, Obrobka_index, Uwagi_index, suma_ilosc_index]
    print("headers", headers_indexes)
    print("\n")
    return List_of_records, headers_indexes


def make_list_of_items_instances(list, indexes):
    items_list = []
    for row in list:
        original_id = row[indexes[2]].lower().strip()

        id = row[indexes[2]].lower().strip().replace("-lustro", "")
        revision = row[indexes[3]]

        id_with_revision = (id + '-' + revision).upper()
        Flag = False

        for instance in items_list:
            if instance.id == id_with_revision:

                if ('lustro' or 'Lustro') in original_id:
                    instance.mirror_quantity = int(
                        instance.mirror_quantity) + int(row[indexes[9]])
                    Flag = True
                    break

                else:
                    instance.quantity = int(
                        instance.quantity) + int(row[indexes[9]])
                    Flag = True
                    break

        if Flag == True:
            continue

        else:
            id_with_revision_instance = items.Item(id_with_revision)

            id_with_revision_instance.revision = row[indexes[3]]
            id_with_revision_instance.description = row[indexes[4]]
            id_with_revision_instance.drawing_Number = row[indexes[5]]
            id_with_revision_instance.material = row[indexes[6]]

            if ('lustro' or 'Lustro') in original_id:
                id_with_revision_instance.mirror_quantity = int(
                    row[indexes[9]])
            else:
                id_with_revision_instance.quantity = int(row[indexes[9]])

            items_list.append(id_with_revision_instance)

    for item in items_list:
        print(item.print_item_values())

    print('items_list: ', items_list)
    return items_list


def Report_to_excel(items_instances_list, order_name, save_in_folder, file_name=""):
    from openpyxl.styles import Font
    # Create a new workbook
    workbook = Workbook()

    # Get the active sheet
    sheet = workbook.active

    # Order number
    sheet.cell(row=1, column=2).value = "Nr zamówienia"
    sheet.cell(row=1, column=4).value = order_name

    sheet.cell(row=2, column=2).value = "Zamawiający"
    sheet.cell(row=2, column=4).value = os.getlogin()
    # Specify row and column indices

    i = 5  # row
    j = 1  # column

    k = 0
    lp = 0
    # pprint.pprint(dictionary_from_excel)
    header_row = 4
    sheet.cell(
        row=header_row, column=1).value = "L.p."
    sheet.cell(
        row=header_row, column=2).value = "ID"
    sheet.cell(
        row=header_row, column=3).value = "Rewizja"
    sheet.cell(
        row=header_row, column=4).value = "Opis"
    sheet.cell(
        row=header_row, column=5).value = "Nr Rysunku"
    sheet.cell(
        row=header_row, column=6).value = "Materiał"
    sheet.cell(
        row=header_row, column=7).value = "Liczba"
    sheet.cell(
        row=header_row, column=8).value = "Liczba luster"
    sheet.cell(
        row=header_row, column=9).value = "Stworzono PDF"

    for item in items_instances_list:

        j = 1
        lp += 1
        sheet.cell(
            row=i, column=j).value = lp
        j = 2
        for item_value in item.print_item_values():
            sheet.cell(
                row=i, column=j).value = item_value

            # Change color for False Cells
            if sheet.cell(row=i, column=j).value == "False":
                for cell in sheet[i]:
                    cell.font = Font(color="FF0000")

            j += 1

        j += 1
        i += 1
        k += 1

    # Enable filtering starting from row 3, from column A to T?Xd
    sheet.auto_filter.ref = f'A4:H{sheet.max_row}'

    # Ustaw szerokość kolumny (np. szerokość kolumny A na 20)
    # Numer kolumny (1 to kolumna A, 2 to kolumna B itd.)
    column_numbers = [1, 2, 3, 4, 5, 6, 7, 8, 9]
    # Pożądana szerokość kolumny (w znakach)
    desired_widths = [5, 15, 10, 40, 60, 20, 10, 15, 15]

    # Ustaw szerokość kolumny
    for column_number, desired_width in zip(column_numbers, desired_widths):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(
            column_number)].width = desired_width

    # Save the workbook
    workbook.save(str(save_in_folder + "/" +
                  order_name + "-" + date_stamp() + ".xlsx"))


def order_items_by_order_number():
    from gui_progress_bar import GUI_make_order
    from gui_progress_bar import Order
    from sql_license import log_in

    # Check License on DataBase
    log_in()

    # Make Class Order
    current_order = Order()
    current_order.user = os.getlogin()
    current_order.date = date_stamp()

    GUI_make_order(current_order)

    # choose xlsm file to prepare order
    open_file = current_order.excel_path

    # choose folder with drawings
    drawings_folder = current_order.drawings_path

    # convert excel to list
    order_name = current_order.order_num

    # choose folder to save files
    current_order.date = date_stamp()
    save_in_folder = current_order.export_path
    save_in_folder = save_in_folder + "/" + order_name + current_order.date
    os.mkdir(save_in_folder)

    Make_list_of_orders = (EXCEL_orders(open_file, order_name))
    orders_list = Make_list_of_orders[0]
    headers_indexes = Make_list_of_orders[1]

    # convert list of orders into dictionary (consolidate lists by id)
    list_of_items_to_order = make_list_of_items_instances(
        orders_list, headers_indexes)

    # open pdf and add info
    # save pdf in folder

    missing_pdf_list = []
    missing_pdf_counter = 0
    pdf_list_to_merge = []

    for item in list_of_items_to_order:
        print('item: ', item)
        id = item.id
        drawing_number = item.drawing_Number.replace(
            "-lustro", "").replace("-Lustro", "")
        print('drawing_number: ', drawing_number)

        amount = item.quantity
        mirror_amount = item.mirror_quantity
        if (int(amount) > 0) and (int(mirror_amount) > 0):
            amount = f'{amount} + {mirror_amount} Lustro '
        elif (int(amount) > 0) and (int(mirror_amount) == 0):
            amount = f'{amount}'
        elif (int(amount) == 0) and (int(mirror_amount) > 0):
            amount = f'{mirror_amount} Lustro'

        else:
            # Rise Error co tu się odjebało!
            pass

        existing_pdf_dir = drawings_folder + "/" + drawing_number + ".pdf"
        print('existing_pdf_dir: ', existing_pdf_dir)
        new_pdf_file_dir = save_in_folder + "/" + drawing_number + \
            "-" + order_name + "-" + current_order.date + ".pdf"
        print('new_file_dir: ', new_pdf_file_dir)
        try:
            # Create new pdf file with new name and read dimensions of page
            pdf = Copy_pdf_to_new_dir(existing_pdf_dir, new_pdf_file_dir)
            pdf_height = int(pdf[1])
            print('pdf_height: ', pdf_height)
            pdf_width = int(pdf[0])
            print('pdf_width: ', pdf_width)

            # Add text box on PDF
            text_box_path = save_in_folder + "/" + id
            create_pdf_with_text_box(
                text_box_path, pdf_height, pdf_width, id, amount, order_name, current_order.date, current_order.user)
            Insert_text_box_on_pdf(new_pdf_file_dir, text_box_path + ".pdf",
                                   new_pdf_file_dir, page_indices="ALL")
            pdf_list_to_merge.append(new_pdf_file_dir)

            # Create QR Code to PNG
            QR_pdf_path = save_in_folder + "/" + id + ".pdf"
            QR_png_path = save_in_folder + "/" + id + ".png"

            QR_png_path = Generate_QR_to_png(id, QR_png_path)
            QR_pdf_path = Convert_QR_PNG_to_PDF(
                QR_png_path, QR_pdf_path, pdf_height, pdf_width)
            Make_QR_stamp_on_pdf(
                new_pdf_file_dir, QR_pdf_path, new_pdf_file_dir)

            # If evrything is okey change PDF Flag True
            item.PDF = True

            try:
                os.remove(QR_png_path)
                os.remove(QR_pdf_path)
            except:
                pass

        except:
            missing_pdf_list.append(drawing_number)
            missing_pdf_counter += 1
            # GUI.Mbox("pdm_search", f"Brak rysunku pliku: {id}", 1)

        # Copy dxf, step to directory
        extensions = [".dxf", ".step", ".stp", ".x_t"]
        for extension in extensions:
            try:
                existing_pdf_dir = drawings_folder + "/" + drawing_number + extension
                print('existing_pdf_dir: ', existing_pdf_dir)
                new_pdf_file_dir = save_in_folder + "/" + drawing_number + \
                    "-" + order_name + "-" + current_order.date + extension
                print('new_file_dir: ', new_pdf_file_dir)
                copy_file(existing_pdf_dir, new_pdf_file_dir)
            except:
                print(f'nie kopiuje {extensions}')
                pass

    # Create one PDF with all drawings
    merge_pdf_path = save_in_folder + "/" + \
        order_name + "-" + current_order.date + ".pdf"
    merge_pdfs(pdf_list_to_merge, merge_pdf_path)

    # Missing drawings information popup
    missing_text = "Zakończono export rysunków do zamówienia" + "\n" + (', '.join(missing_pdf_list)) + "\n" + \
        f"Liczba brakujących pdfów: {missing_pdf_counter}"
    GUI.Mbox("pdm_search", missing_text, 1)

    """# Missing drawings txt report
    report_file_name = save_in_folder + "/" + "000_Missing_Drawings" + \
        "-" + order_name + "-" + current_order.date + ".txt"
    with open(report_file_name, 'w', encoding="utf-8") as report:
        if len(missing_pdf_list) != 0:
            report.write("Lista brakujących rysunków z zamówienia.\n")
            report.write('\n'.join(missing_pdf_list))

        elif len(missing_pdf_list) == 0:
            report.write(
                f'{current_order.date}, Wszystkie rysunki zostały wygenerowane. Nie odnotowano braków.\n\n')

            i = 0
            for item in list_of_items_to_order:
                i += 1
                values = []
                for value in item.print_item_values():
                    values.append(str(value))

                report.write(f'{i}.\t')
                report.write(", ".join(values))
                report.write("\n")"""

    # Make report from parts ordered
    Report_to_excel(list_of_items_to_order, order_name, save_in_folder)

    # SQL check login and license

    user_id = SQL.read_db(
        f"SELECT user_id FROM Users WHERE login = '{current_order.user}';")
    user_id = int(user_id[0][0])
    print('user_id: ', user_id)

    company_id = SQL.read_db(
        f"SELECT company_id FROM Users WHERE login = '{current_order.user}';")
    company_id = company_id[0][0]
    print('company_id: ', company_id)

    license_id = SQL.read_db(
        f"SELECT license_id FROM Licenses WHERE company_id = '{company_id}';")
    license_id = license_id[0][0]
    print('license_id: ', license_id)

    license_check = SQL.read_db(
        f"SELECT license_number FROM Licenses WHERE license_id = '{license_id}';")
    license_check = bool(license_check[0][0])

    if license_check == True:

        insert_sql_order = ("INSERT INTO Orders (order_id, order_number, user_id, order_project)" +
                            f"VALUES (default,'{order_name}', '{user_id}', '0000');")
        SQL.input_db(insert_sql_order)
    else:
        sys.exit()


# TESTS
"""pdf = Input_data_on_pdf("test.pdf", "test_new.pdf", "123")
pdf_height = int(pdf[1])
print('pdf_height: ', pdf_height)
pdf_width = int(pdf[0])
print('pdf_width: ', pdf_width)

stamp_pdf_path = Convert_PNG_to_PDF(id, pdf_height, pdf_width)
Make_stamp_on_pdf("test_new.pdf", stamp_pdf_path, "test_new.pdf")
"""
