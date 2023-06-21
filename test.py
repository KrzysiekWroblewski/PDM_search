import openpyxl
from excel_raport_generator import Report
from List_sorting import Sort
from gui import GUI


def write_excel_BOM_to_list(open_file: str) -> list[str]:
    """Def open active sheet of work book and write rows to lists

    Args:
        open_file (str): Path to file which is rewrite to list

    Returns:
        list[str]: strings in list are encoded to bytes
    """

    wrkbk = openpyxl.load_workbook(open_file)
    sh = wrkbk.active

    list_of_records = []

    for i in range(5, sh.max_row+1):
        details_of_record = []
        for j in range(1, sh.max_column+1):
            details_of_record.append(
                str(sh.cell(row=i, column=j).value).encode("utf-8"))
        list_of_records.append(details_of_record)
    print('list_of_records: ', list_of_records)

    return list_of_records


def find_items_without_value_in_cell(list_of_items: list[str], searched_column: int):
    list_of_found_items = []
    for item in list_of_items:

        if "none" in item[searched_column].decode("utf-8").lower():
            list_of_found_items.append(item)
            print("i Found unordered item!")
            continue

    print('list_of_found_items: ', list_of_found_items)
    return list_of_found_items


def znajdz_indeks(lista, szukany_element):
    for indeks, element in enumerate(lista[0]):
        if element.decode("utf-8") in szukany_element:
            return indeks
    return -1  # zwracamy -1, jeśli element nie został znaleziony


"""List_of_items = write_excel_BOM_to_list(open_file)
List_of_items_with_missing_values = find_items_without_value_in_cell(
    List_of_items, 19)
sorted_list_of_items_with_missing_values = Sort.list_sort_by_index(
    List_of_items_with_missing_values.copy(), 5)

print(sorted_list_of_items_with_missing_values)
Report.generate_report_txt(sorted_list_of_items_with_missing_values, "name_of_filestr", "search_Path_sld_prt: bytes".encode(
    "utf-8"), "search_Path_sld_prt_revision: bytes".encode("utf-8"), [0, 1, 2, 3, 4, 5, 6, 7, 8])"""
