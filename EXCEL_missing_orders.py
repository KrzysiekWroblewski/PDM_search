import openpyxl
from EXCEL_report_generator import Report
from List_sorting import Sort
from gui import GUI
import os
from EXCEL_big_data import Report_missing_orders_to_excel


def Return_file_name(open_file: str) -> str:
    """Function check selected file, returns name of file without extension

    Args:
        open_file (str): filepath in str

    Returns:
        _type_: extension e.g. "my_file"
    """
    file_name = os.path.basename(open_file)
    name_without_extension = file_name.split(".")
    return name_without_extension[0]

# def check selected file, returns extension of file and file_path


def Return_file_type(open_file: str) -> str:
    """Function check selected file, returns extension of file
    Args:
        open_file (str):    

    Returns:
        str: extension of te file e.g. "txt"
    """
    file_name = os.path.basename(open_file)
    name_without_extension = file_name.split(".")
    return name_without_extension[1]


def List_of_excel_files_from_txt(open_file) -> list[str]:
    # select txt file with file paths of many files each in new line

    encode_list = ["latin-1", "utf-16-le", "utf-8", "CP-1250"]

    i = 0
    list_of_files = []
    for _ in encode_list:
        try:
            # merging text function
            with open(open_file, encoding=encode_list[i], mode='r') as file:
                # print('open_file: ', open_file)
                for line in file:
                    line = line.strip()
                    list_of_files.append(line)
            break

        except:
            print("Decoding", encode_list[i], "do not work")
            i = i + 1

    return list_of_files


def Merge_search_path(list, index1, index2=None):

    search_Path_sld_prt = "".encode("utf-8")
    search_Path_sld_prt_revision = "".encode("utf-8")

    for row in list:
        search_Path_sld_prt += row[index1] + "|".encode('utf-8')
        if index2 != None:
            search_Path_sld_prt_revision += row[index1] + \
                "-".encode('utf-8') + row[index2] + "|".encode('utf-8')

    # search_Path_sld_prt = search_Path_sld_prt.encode('utf-8')
    # search_Path_sld_prt_revision = search_Path_sld_prt_revision.encode('utf-8')
    return search_Path_sld_prt, search_Path_sld_prt_revision


def EXCEL_missing_orders(open_file: str, sort_by: int):
    List_of_items = Write_excel_BOM_to_list(open_file, start_row=5)

    # find column with header "Numer zamówienia"
    header_title = "Numer zamówienia"
    column_index = Find_column_index_by_string(List_of_items, header_title)

    # find rows with None in column
    List_of_records = Find_items_without_value_in_cell(
        List_of_items, column_index)

    sorted_list_of_items_with_missing_values = Sort.list_sort_by_index(
        List_of_records.copy(), sort_by)

    return sorted_list_of_items_with_missing_values


def Write_excel_BOM_to_list(open_file: str, start_row: int) -> list[str]:
    """Def open active sheet of work book and write rows to lists

    Args:
        open_file (str): Path to file which is rewrite to list

    Returns:
        list[str]: strings in list are encoded to bytes
    """

    wrkbk = openpyxl.load_workbook(open_file)
    sh = wrkbk.active

    list_of_records = []

    for i in range(start_row, sh.max_row+1):
        details_of_record = []
        for j in range(1, sh.max_column+1):
            details_of_record.append(
                str(sh.cell(row=i, column=j).value).replace("\n", "").encode("utf-8"))
        list_of_records.append(details_of_record)
    # print('list_of_records: ', list_of_records)

    return list_of_records


def Find_items_without_value_in_cell(list_of_items: list[str], searched_column: int):
    list_of_found_items = []
    for item in list_of_items:

        if "none" in item[searched_column].decode("utf-8").lower():
            list_of_found_items.append(item)
            # print("i Found unordered item!")
            continue

    # print('list_of_found_items: ', list_of_found_items)
    return list_of_found_items


def Find_column_index_by_string(lista, szukany_element):
    for indeks, element in enumerate(lista[0]):
        if element.decode("utf-8") in szukany_element:
            return indeks
    return -1  # zwracamy -1, jeśli element nie został znaleziony


def mega_def(indexes_to_report, sort_by):
    # checking type of selected file txt or xlsx
    open_file = GUI.select_file()
    file_extension = Return_file_type(open_file)
    file_name_without_extension = Return_file_name(open_file)
    list_of_excels_with_missing_orders = []
    list_search_Path_sld_prt = []
    list_search_Path_sld_prt_revision = []

    # if user will select txt file with list of many excels, all of them will be reported
    if file_extension == "txt":
        list_of_excels_to_report = List_of_excel_files_from_txt(open_file)
        # print('list_to_report: ', list_of_excels_to_report)
        i = 0
        for list in list_of_excels_to_report:

            list_of_excels_with_missing_orders.append(
                EXCEL_missing_orders(list, sort_by))

            search_Path = Merge_search_path(
                list_of_excels_with_missing_orders[i], 6, 7)

            list_search_Path_sld_prt.append(search_Path[0])

            list_search_Path_sld_prt_revision.append(search_Path[1])

            i += 1

    # if user will select excel file, only this file will be reported
    elif file_extension == "xlsx" or "xlsm":

        list_of_excels_to_report = [open_file]
        list_of_excels_with_missing_orders.append(
            EXCEL_missing_orders(open_file, sort_by))

        search_Path = Merge_search_path(
            list_of_excels_with_missing_orders[0], 6, 7)

        list_search_Path_sld_prt.append(search_Path[0])

        list_search_Path_sld_prt_revision.append(search_Path[1])

    else:
        print("Zły format pliku")
        GUI.Mbox("Wyszukiwarka PDM via CSV",
                 "Zły format pliku", 1)

    GUI.Mbox("Wyszukiwarka PDM via CSV",
             "Wybierz folder gdzie zapisać raport braków", 1)

    date_stamp = Report.date_stamp()
    file = ((GUI.select_folder()) + "/" + date_stamp +
            "_" + file_name_without_extension + "_report.txt")

    # Write report to .txt
    Report.Report_list_to_txt("Lista brakujących zamowien!", file, list_of_excels_to_report,
                              list_of_excels_with_missing_orders, date_stamp, indexes_to_report, list_search_Path_sld_prt, list_search_Path_sld_prt_revision)

    # Write report to .slsx
    Report_missing_orders_to_excel("Lista brakujących zamowien!", file_name_without_extension, file, list_of_excels_to_report,
                                   list_of_excels_with_missing_orders, indexes_to_report, list_search_Path_sld_prt, list_search_Path_sld_prt_revision)
