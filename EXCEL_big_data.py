from openpyxl import Workbook
from EXCEL_report_generator import Report


def Report_missing_orders_to_excel(title_message="", file_name_without_extension="", file_directory="", list_of_excels_to_report="", list_of_excels_with_missing_orders="", indexes_to_report="", list_search_Path_sld_prt="", list_search_Path_sld_prt_revision=""):

    title_message = "To jest lista elementów, które nie zostały jeszcze zamówione"
    headers = ["Lp", "data", ".PDF", ".DXF", ".STEP", "Konstruktor", "ID. Części", "Rewizja", "Opis",
               "Nr. Części", "Nr. Kat.", "Materiał", "Obróbka", "Obróbka cieplna",	"Powłoka", "komentarz", "Uwagi", "Projekt"]

    # Create a new workbook
    workbook = Workbook()

    # Get the active sheet
    sheet = workbook.active

    # Specify row and column indices
    row_index = 1
    column_index = 1

    # Write a string into the specified cell
    title_message = "To jest lista elementów, które nie zostały jeszcze zamówione"
    date_stamp = Report.date_stamp()

    sheet.cell(row=1, column=1).value = title_message
    sheet.cell(row=2, column=1).value = date_stamp

    i = 3
    j = 1
    for title in headers:
        sheet.cell(row=i, column=j).value = title
        j += 1

    i = 4  # row
    j = 1  # column

    k = 0

    for list in list_of_excels_with_missing_orders:

        for row in list:

            j = 1
            for index in indexes_to_report:
                sheet.cell(row=i, column=j).value = row[index]
                j += 1

            # add name of file from where is the missing part
            sheet.cell(row=i, column=j).value = list_of_excels_to_report[k].split(
                "\\")[-1].split(".")[-2]  # xD bo ja bede pamiętał co tu zrobiłem

            j += 1
            i += 1
        k += 1

    # Enable filtering starting from row 3, from column A to T?Xd
    sheet.auto_filter.ref = f'A3:T{sheet.max_row}'
    # Save the workbook
    print("file dir: ", file_directory)
    workbook.save(str(file_directory + "_report.xlsx"))
