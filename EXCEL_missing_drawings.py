import openpyxl
from EXCEL_report_generator import Report
from gui import GUI
import openpyxl
from EXCEL_report_generator import Report
from gui import GUI
from EXCEL_missing_orders import Return_file_type
from EXCEL_missing_orders import Return_file_name

# open_file = "sample.xlsx"


def missing_drawings_from_excel():

    open_file = GUI.select_file()
    file_extension = Return_file_type(open_file)
    file_name_without_extension = Return_file_name(open_file)
    list_of_excels_to_report = [open_file]
    indexes_to_report = [0, 1, 2, 3, 4]

    xl_part_column = 7
    csv_sep = " | "

    wrkbk = openpyxl.load_workbook(open_file)
    sh = wrkbk.active

    search_Path_sld_prt = ""
    search_Path_sld_prt_revision = ""
    number_of_files_merged = 0

    list_of_records = []

    for j in range(6, sh.max_row+1):
        cell_obj = sh.cell(row=j, column=xl_part_column)
        cell_revision = sh.cell(row=j, column=xl_part_column+1)

        # skip line if par is mirror
        try:
            if "lustro" in sh.cell(row=j, column=xl_part_column).value.lower():
                continue
        except:
            continue

        # Checking if  are some missing drawings
        for i in range(3, 6):
            if sh.cell(row=j, column=i).value == None:
                drawings = False
                break
            else:
                drawings = True

        if drawings == False:
            list_of_records.append(
                [str(sh.cell(row=j, column=1).value).encode("utf-8"), str(sh.cell(row=j, column=6).value).encode("utf-8"), str(sh.cell(row=j, column=xl_part_column).value).encode("utf-8"), str(sh.cell(row=j, column=xl_part_column+1).value).encode("utf-8"), str(sh.cell(row=j, column=xl_part_column+2).value).encode("utf-8")])

        if drawings == False:
            search_Path_sld_prt = search_Path_sld_prt + \
                str(cell_obj.value) + csv_sep

            number_of_files_merged = number_of_files_merged + 1

        if drawings == False:
            search_Path_sld_prt_revision = search_Path_sld_prt_revision + \
                str(cell_obj.value) + "-" + str(cell_revision.value) + csv_sep

    try:
        if search_Path_sld_prt[-3:] == " | ":
            search_Path_sld_prt = search_Path_sld_prt[:-3]
            search_Path_sld_prt_revision = search_Path_sld_prt_revision[:-3]
    except:
        pass

    search_Path_sld_prt = [search_Path_sld_prt.encode(
        'utf-8')]
    # .decode('ascii', 'ignore')

    search_Path_sld_prt_revision = [search_Path_sld_prt_revision.encode(
        'utf-8')]
    # .decode('ascii', 'ignore')

    print(search_Path_sld_prt)
    print(search_Path_sld_prt_revision)
    print(number_of_files_merged)

    # Print Report to txt of missing drawings
    date_stamp = Report.date_stamp()
    file = ((GUI.select_folder()) + "/" + date_stamp +
            "_" + file_name_without_extension + "_report.txt")

    list_of_excels_with_missing_orders = [list_of_records]
    print('list_of_excels_with_missing_orders: ',
          list_of_excels_with_missing_orders)

    Report.Report_list_to_txt("Lista brakujacych Rysunkow!", file, list_of_excels_to_report,
                              list_of_excels_with_missing_orders, date_stamp, indexes_to_report, search_Path_sld_prt, search_Path_sld_prt_revision)

    # Window with message
    Message_01 = str("Source file path: " + str(open_file) +
                     str("\n") + str("\n") + "Number of missing files: " + str(number_of_files_merged))
    GUI.Mbox("Wyszukiwarka plikow PDM", Message_01, 1)

    return True
