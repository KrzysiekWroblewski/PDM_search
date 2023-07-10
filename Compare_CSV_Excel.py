from numpy import column_stack
from EXCEL_methods import excel
from gui import GUI
import csv
from openpyxl import Workbook
from EXCEL_report_generator import Report
import pprint


def Return_file_name(open_file: str) -> str:
    """Function check selected file, returns name of file without extension

    Args:
        open_file (str): filepath in str

    Returns:
        _type_: extension e.g. "my_file"
    """
    file_name = os.path.basename(open_file)
    name_without_extension = file_name.split(".")
    return name_without_extension[0]


def write_list_to_dict(list, indexes):
    row_list_len = len(list[0])
    dict1 = {}
    dict1_list = []
    counter = 0
    for row in list:
        counter += 1
        # Skipp no PRT-Files
        if "PRT-" not in row[indexes[0]]:
            continue

        # Sometimes CSV from PDM have wrong ammount of columns, we skipp them and give message about error
        if len(row) < row_list_len:
            GUI.Mbox(
                "PDM_Search", f"Problem with line {counter }, wrong value of columns", 1)
            continue

        # jeśli już jest id w słowniku to sumuj te wartości
        if row[indexes[0]] in dict1.keys():
            dict1[row[indexes[0]]
                  ][-1] = int(dict1[row[indexes[0]]][-1]) + int(row[indexes[-1]])
            print("powtarza się dict")

        # jeśli nie ma id w słowniku, utwórz nowy klucz i wartości
        else:

            record = {row[indexes[0]]: [row[indexes[0]],
                                        row[indexes[1]], row[indexes[2]], row[indexes[3]]]}
            print(record)
            dict1.update(record)

    dict1_list.append(dict1)

    # print('dict1_list: ', dict1_list)
    # print("\n")s
    return dict1_list


def Find_missing_items(main_dictionary, search_in_dict):
    main_list_keys = []
    search_in_list_keys = []
    miss_list = []

    # stworz liste kluczy
    for item in main_dictionary[0].keys():
        main_list_keys.append(item)
    # print('excel_keys: ', main_list_keys)

    # stworz liste kluczy
    for item in search_in_dict[0].keys():
        search_in_list_keys.append(item)
    # print('csv_keys: ', search_in_list_keys)

    # Jesli nie ma klucza w drugiej liscie dodaj go do listy brakujacych elementow
    for item in main_list_keys:

        if item in search_in_list_keys:
            id_main = main_dictionary[0][item][0]
            revision_main = main_dictionary[0][item][1]
            description_main = main_dictionary[0][item][2]
            value_main = int(main_dictionary[0][item][-1])

            value_secondary = int(search_in_dict[0][item][-1])

            # pomijamy jelsi liczba sie zgadza na obu listach
            if value_main == value_secondary:
                # print("Jest tyle samo elementów")
                continue

            # tutaj oblcziamy o ile za duzo elementow zamowionych jest w excel
            elif value_main > value_secondary:
                a = [id_main, revision_main, description_main,
                     value_main - value_secondary]
                miss_list.append(a)
                # print("ilosc_main_dict > ilosc_secondary_dict")

            # tutaj oblcziamy o ile za za mało zostało wpisanych elementow do excela
            elif value_main < value_secondary:
                b = [id_main, revision_main, description_main,
                     value_main - value_secondary]
                miss_list.append(b)
                # print("ilosc_main_dict < ilosc_secondary_dict")

            else:
                print("Stalo sie else")

        elif item not in search_in_list_keys:
            id_main = main_dictionary[0][item][0]
            revision_main = main_dictionary[0][item][1]
            description_main = main_dictionary[0][item][2]
            value_main = int(main_dictionary[0][item][-1])

            b = [id_main, revision_main, description_main, value_main]
            miss_list.append(b)

    for item in search_in_list_keys:
        if item not in main_list_keys:
            name_secondary = search_in_dict[0][item][0]
            revision_secondary = search_in_dict[0][item][1]
            description_secondary = search_in_dict[0][item][2]
            value_secondary = int(search_in_dict[0][item][-1])

            b = [name_secondary, revision_secondary,
                 description_secondary, - value_secondary]
            miss_list.append(b)

    print("\n")
    # print("miss_list: ", miss_list)

    return miss_list


def make_dictionary_from_excel():
    open_file = GUI.select_file()
    start_row = 5
    excel_list = excel.Write_excel_BOM_to_list(open_file, start_row)

    id_index = excel.Excel_Find_column_index_by_string_in_list(
        excel_list, "ID. Części")
    Lp_index = excel.Excel_Find_column_index_by_string_in_list(
        excel_list, "Lp")
    Konstruktor_index = excel.Excel_Find_column_index_by_string_in_list(
        excel_list, "Konstruktor")
    Rewizja_index = excel.Excel_Find_column_index_by_string_in_list(
        excel_list, "Rewizja")
    Opis_index = excel.Excel_Find_column_index_by_string_in_list(
        excel_list, "Opis")
    do_Zamowienia_index = excel.Excel_Find_column_index_by_string_in_list(
        excel_list, "Do Zamówienia")

    columns = [id_index, Rewizja_index, Opis_index, do_Zamowienia_index]
    excel_list.pop(0)

    dictionary_from_excel = write_list_to_dict(excel_list, columns)

    return dictionary_from_excel


def make_dictionary_from_CSV(csv_delimiter):
    open_file = GUI.select_file()

    with open(open_file, encoding="utf-8", mode='r') as csvfile:
        csv_reader = csv.reader(csvfile, delimiter=str(csv_delimiter))
        list = []
        for line in csv_reader:
            list.append(line)
            if list[-1] == []:
                list.pop(-1)

        id_index = excel.Excel_Find_column_index_by_string_in_list(
            list, "ID. Części")
        Rewizja_index = excel.Excel_Find_column_index_by_string_in_list(
            list, "Rewizja")
        Opis_index = excel.Excel_Find_column_index_by_string_in_list(
            list, "Opis")
        do_Zamowienia_index = excel.Excel_Find_column_index_by_string_in_list(
            list, "Liczba odniesień")

        columns = [id_index, Rewizja_index,
                   Opis_index, do_Zamowienia_index]
        list.pop(0)

        dictionary_from_CSV = write_list_to_dict(list, columns)

    return dictionary_from_CSV


def Report_to_excel(dictionary_from_excel, file_name=""):
    # Create a new workbook
    workbook = Workbook()

    # Get the active sheet
    sheet = workbook.active

    # Specify row and column indices

    i = 4  # row
    j = 1  # column

    k = 0

    for item in dictionary_from_excel:

        j = 1
        for index in [0, 1, 2, 3]:
            sheet.cell(row=i, column=j).value = item[index]
            j += 1

        j += 1
        i += 1
        k += 1

    # Enable filtering starting from row 3, from column A to T?Xd
    sheet.auto_filter.ref = f'A3:T{sheet.max_row}'
    # Save the workbook
    date_stamp = Report.date_stamp()
    select_folder = GUI.select_folder()
    workbook.save(str(select_folder + "/" + date_stamp + file_name +
                  "Braki z dictionary_report.xlsx"))


def Compare_CSV_to_Excel():
    excel_dict = make_dictionary_from_excel()
    pprint.pprint(excel_dict)

    print("\n")

    csv_dict = make_dictionary_from_CSV(csv_delimiter=";")
    pprint.pprint(csv_dict)

    print("\n")

    compared_dict = Find_missing_items(excel_dict, csv_dict)
    pprint.pprint(compared_dict)

    Report_to_excel(compared_dict)
