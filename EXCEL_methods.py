import openpyxl


class excel():

    @staticmethod
    def Write_excel_BOM_to_list(open_file: str, start_row: int) -> list[str]:
        """Def open active sheet of work book and write rows to lists

        Args:
            open_file (str): Path to file which is rewrite to list

        Returns:
            list[str]: strings in list are encoded to bytes
        """

        wrkbk = openpyxl.load_workbook(open_file)
        sh = wrkbk.active

        list_of_records = []

        for i in range(start_row, sh.max_row+1):
            details_of_record = []
            for j in range(1, sh.max_column+1):
                details_of_record.append(
                    str(sh.cell(row=i, column=j).value).replace("\n", ""))
            list_of_records.append(details_of_record)
        # print('list_of_records: ', list_of_records)

        return list_of_records

    @staticmethod
    def Excel_Find_items_without_value_in_cell(list_of_items: list[str], searched_column: int):
        list_of_found_items = []
        for item in list_of_items:

            if "none" in item[searched_column].decode("utf-8").lower():
                list_of_found_items.append(item)
                # print("i Found unordered item!")
                continue

        # print('list_of_found_items: ', list_of_found_items)
        return list_of_found_items

    @staticmethod
    def Excel_Find_column_index_by_string_in_list(lista, header_name):
        for indeks, element in enumerate(lista[0]):
            if element in header_name:
                return indeks
        return -1  # zwracamy -1, jeśli element nie został znaleziony
