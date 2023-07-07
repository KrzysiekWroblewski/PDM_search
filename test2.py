from numpy import column_stack
from EXCEL_methods import excel
from gui import GUI
import csv
from openpyxl import Workbook
from EXCEL_report_generator import Report


def write_list_to_dict(list, indexes):
    dict1 = {}
    dict1_list = []
    for row in list:

        # jeśli już jest id w słowniku to sumuj te wartości
        if row[indexes[0]] in dict1.keys():
            # print("o panie: ", dict1[(item[indexes])][-1])a
            # print("jaki  to prt: ", row[indexes[0]])
            # print("nr kolumny id: ", indexes[0])
            # print("dict1[(item[indexes])][-1]", dict1[row[indexes[0]]])

            # sumuj ilsoci do zamowienia dla danego id
            dict1[row[indexes[0]]
                  ][-1] = int(dict1[row[indexes[0]]][-1]) + int(row[indexes[-1]])

        # jeśli nie ma id w słowniku, utwórz nowy klucz i wartości
        else:

            rekord = {row[indexes[0]]: [row[indexes[0]], row[indexes[1]],
                                        row[indexes[2]], row[indexes[3]]]}
            dict1.update(rekord)

    dict1_list.append(dict1)

    # print('dict1_list: ', dict1_list)
    # print("\n")s
    return dict1_list


def Find_missing_items(main_dictionary, search_in_dict):
    main_list_keys = []
    search_in_list_keys = []
    miss_list = []

    # stworz liste kluczy
    for item in main_dictionary[0].keys():
        main_list_keys.append(item)
    # print('excel_keys: ', main_list_keys)

    # stworz liste kluczy
    for item in search_in_dict[0].keys():
        search_in_list_keys.append(item)
    # print('csv_keys: ', search_in_list_keys)

    # Jesli nie ma klucza w drugiej liscie dodaj go do listy brakujacych elementow
    for item in main_list_keys:

        if item in search_in_list_keys:
            nazwa_main = main_dictionary[0][item][0]
            ilosc_main_dict = int(main_dictionary[0][item][-1])
            ilosc_secondary_dict = int(search_in_dict[0][item][-1])

            # pomijamy jelsi liczba sie zgadza na obu listach
            if ilosc_main_dict == ilosc_secondary_dict:
                print("Jest tyle samo elementów")
                continue

            # tutaj oblcziamy o ile za duzo elementow zamowionych jest w excel
            elif ilosc_main_dict > ilosc_secondary_dict:
                a = [nazwa_main, ilosc_main_dict - ilosc_secondary_dict]
                miss_list.append(a)
                print("ilosc_main_dict > ilosc_secondary_dict")

            # tutaj oblcziamy o ile za za mało zostało wpisanych elementow do excela
            elif ilosc_main_dict < ilosc_secondary_dict:
                b = [nazwa_main, ilosc_main_dict - ilosc_secondary_dict]
                miss_list.append(b)
                print("ilosc_main_dict < ilosc_secondary_dict")

            else:
                print("Stalo sie else")

        elif item not in search_in_list_keys:
            nazwa_main = main_dictionary[0][item][0]
            ilosc_main_dict = int(main_dictionary[0][item][-1])
            b = [nazwa_main, ilosc_main_dict]
            miss_list.append(b)

    for item in search_in_list_keys:
        if item not in main_list_keys:
            nazwa_secondary = search_in_dict[0][item][0]
            ilosc_secondary_dict = int(search_in_dict[0][item][-1])
            b = [nazwa_secondary, -ilosc_secondary_dict]
            miss_list.append(b)

    print("\n")
    # print("miss_list: ", miss_list)

    return miss_list


def make_dictionary_from_excel():
    open_file = GUI.select_file()
    start_row = 5
    excel_list = excel.Write_excel_BOM_to_list(open_file, start_row)

    id_index = excel.Excel_Find_column_index_by_string_in_list(
        excel_list, "ID. Części")
    Lp_index = excel.Excel_Find_column_index_by_string_in_list(
        excel_list, "Lp")
    Konstruktor_index = excel.Excel_Find_column_index_by_string_in_list(
        excel_list, "Konstruktor")
    Rewizja_index = excel.Excel_Find_column_index_by_string_in_list(
        excel_list, "Rewizja")
    Opis_index = excel.Excel_Find_column_index_by_string_in_list(
        excel_list, "Opis")
    do_Zamowienia_index = excel.Excel_Find_column_index_by_string_in_list(
        excel_list, "Do Zamówienia")

    columns = [id_index, Rewizja_index, Opis_index, do_Zamowienia_index]
    excel_list.pop(0)

    dictionary_from_excel = write_list_to_dict(excel_list, columns)
    # print("\n")
    # print('id_index: ', id_index)
    # print("dictionary_from_excel: ", dictionary_from_excel)
    return dictionary_from_excel


def make_dictionary_from_CSV(csv_delimiter):
    open_file = GUI.select_file()
    with open(open_file, encoding="utf-8", mode='r') as csvfile:
        csv_reader = csv.reader(csvfile, delimiter=str(csv_delimiter))
        list = []
        for line in csv_reader:
            list.append(line)
            if list[-1] == []:
                list.pop(-1)

    id_index = excel.Excel_Find_column_index_by_string_in_list(
        list, "ID. Części")
    Rewizja_index = excel.Excel_Find_column_index_by_string_in_list(
        list, "Rewizja")
    Opis_index = excel.Excel_Find_column_index_by_string_in_list(
        list, "Opis")
    do_Zamowienia_index = excel.Excel_Find_column_index_by_string_in_list(
        list, "Liczba odniesień")

    columns = [id_index, Rewizja_index, Opis_index, do_Zamowienia_index]
    list.pop(0)
    # print("\n")
    # print('columns_indexes: ', columns)

    dictionary_from_excel = write_list_to_dict(list, columns)

    # print("\n")
    # print('columns_indexes: ', columns)
    # print(dictionary_from_excel)

    return dictionary_from_excel


def Report_to_excel(dictionary_from_excel):
    # Create a new workbook
    workbook = Workbook()

    # Get the active sheet
    sheet = workbook.active

    # Specify row and column indices

    i = 4  # row
    j = 1  # column

    k = 0

    for item in dictionary_from_excel:

        j = 1
        for index in range(0, 2):
            sheet.cell(row=i, column=j).value = item[index]
            j += 1

        j += 1
        i += 1
        k += 1

    # Enable filtering starting from row 3, from column A to T?Xd
    sheet.auto_filter.ref = f'A3:T{sheet.max_row}'
    # Save the workbook
    workbook.save(str("Braki z dictionary_report.xlsx"))


excel_dict = make_dictionary_from_excel()
csv_list = make_dictionary_from_CSV(csv_delimiter=";")
a = Find_missing_items(excel_dict, csv_list)
Report_to_excel(a)
